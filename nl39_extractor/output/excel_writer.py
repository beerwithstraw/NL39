"""
Excel Writer — generates the multi-sheet Excel output with formatting.

Source: approach document Section 11
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config.settings import (
    MASTER_COLUMNS,
    EXTRACTOR_VERSION,
    NUMBER_FORMAT,
    LOW_CONFIDENCE_FILL_COLOR,
    company_key_to_pascal
)
from config.row_registry import ROW_ORDER, ROW_DISPLAY_NAMES
from config.lob_registry import LOB_ORDER, LOB_DISPLAY_NAMES
from config.company_metadata import get_metadata
from config.lob_metadata import get_lob_particulars, get_grouped_lob
from extractor.models import CompanyExtract

logger = logging.getLogger(__name__)

# Style definitions
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
_META_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_YELLOW_FILL = PatternFill(start_color=LOW_CONFIDENCE_FILL_COLOR, end_color=LOW_CONFIDENCE_FILL_COLOR, fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

def _year_code_to_fy_end(year_code: str) -> str:
    """Convert year code to FY end year.  '20242025' -> '2025', '202425' -> '2025'."""
    s = str(year_code).strip()
    if len(s) == 8:
        return s[4:]       # 20242025 -> 2025
    if len(s) == 6:
        return f"20{s[4:]}"  # 202425 -> 2025
    return s


# Columns that hold numeric metric values (for number formatting)
_METRIC_COLUMNS = {c for c in MASTER_COLUMNS if c.lower() in ROW_ORDER}


def _write_master_data(ws, extractions: List[CompanyExtract], existing_rows: Optional[List[list]] = None, year_selection: str = "both"):
    """Writes the Master_Data sheet with the NL-39 column structure."""
    # 1. Header Row
    for col_idx, col_name in enumerate(MASTER_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER_ALIGN

    ws.freeze_panes = "A2"

    current_row = 2

    # 2a. Write preserved rows from previous runs
    if existing_rows:
        for row_data in existing_rows:
            for col_idx, val in enumerate(row_data, 1):
                if col_idx > len(MASTER_COLUMNS):
                    break
                cell = ws.cell(row=current_row, column=col_idx, value=val)
                if MASTER_COLUMNS[col_idx - 1] in _METRIC_COLUMNS:
                    cell.number_format = NUMBER_FORMAT
            current_row += 1

    # 2b. Data Rows from new extractions
    # NL-39 has no prior year — only current_year is populated.
    for extract in extractions:
        meta = get_metadata(extract.company_key)
        period_data = extract.current_year
        if not period_data:
            continue

        for lob in LOB_ORDER:
            if lob not in period_data.data:
                continue

            for q_status, p_key in [("For the Quarter", "qtr"), ("Up to the Quarter", "ytd")]:
                lob_data = period_data.data.get(lob, {})

                # Build metadata lookup for non-metric columns
                metadata = {
                    "LOB_PARTICULARS": get_lob_particulars(lob),
                    "Grouped_LOB": get_grouped_lob(lob),
                    "Company_Name": meta["company_name"],
                    "Company": meta["sorted_company"],
                    "NL": extract.form_type,
                    "Quarter": extract.quarter,
                    "Year": _year_code_to_fy_end(extract.year),
                    "Quarter_Info": q_status,
                    "Sector": meta["sector"],
                    "Industry_Competitors": meta["competitors"],
                    "GI_Companies": "GI Company",
                    "Source_File": extract.source_file,
                }

                row_values = []
                for col_name in MASTER_COLUMNS:
                    if col_name in metadata:
                        row_values.append(metadata[col_name])
                    elif col_name.lower() in ROW_ORDER:
                        val = lob_data.get(col_name.lower(), {}).get(p_key)
                        row_values.append(val)
                    else:
                        row_values.append(None)

                # Write to sheet
                for col_idx, val in enumerate(row_values, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=val)
                    if MASTER_COLUMNS[col_idx - 1] in _METRIC_COLUMNS:
                        cell.number_format = NUMBER_FORMAT

                current_row += 1

def _write_verification_sheet(ws, extract: CompanyExtract):
    """
    Writes per-company verification sheet matching the PDF layout:
      - Rows  = LOBs  (Fire, Marine Cargo, ...)
      - Cols  = Metrics (count buckets, amount buckets, totals)
      - Table 1: For the Quarter (qtr values)
      - Table 2: Upto the Quarter (ytd values)
    """
    ws.cell(row=1, column=1, value=f"VERIFICATION SHEET: {extract.company_name}").font = Font(bold=True, size=14)
    ws.cell(row=2, column=1, value=f"Quarter: {extract.quarter} | Year: {extract.year} | Source: {extract.source_file}")

    if not extract.current_year:
        ws.cell(row=4, column=1, value="No data extracted.").font = Font(italic=True)
        return

    active_lobs = [lob for lob in LOB_ORDER if lob in extract.current_year.data]

    # TABLE 1: For the Quarter
    qtr_start = 4
    _write_pdf_table(ws, extract.current_year, active_lobs, start_row=qtr_start,
                     title="TABLE 1: For the Quarter", p_key="qtr")

    # TABLE 2: Upto the Quarter (gap of 3 rows)
    ytd_start = qtr_start + 2 + len(active_lobs) + 3
    _write_pdf_table(ws, extract.current_year, active_lobs, start_row=ytd_start,
                     title="TABLE 2: Upto the Quarter", p_key="ytd")


def _write_pdf_table(ws, period_data, active_lobs: list, start_row: int, title: str, p_key: str):
    """
    Write one NL-39 table (QTR or YTD) in PDF orientation:
      - Row 0: title
      - Row 1: metric column headers
      - Rows 2+: one row per LOB, metric values in columns
    """
    if not period_data:
        ws.cell(row=start_row, column=1, value=f"{title} (Data Not Found)").font = Font(italic=True)
        return

    # Title row
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    ws.merge_cells(
        start_row=start_row, start_column=1,
        end_row=start_row, end_column=1 + len(ROW_ORDER)
    )
    title_cell.fill = _HEADER_FILL
    title_cell.font = Font(bold=True, color="FFFFFF", size=11)

    # Header row: "Line of Business" | metric1 | metric2 | ...
    h_row = start_row + 1
    lob_header = ws.cell(row=h_row, column=1, value="Line of Business")
    lob_header.fill = _META_FILL
    lob_header.font = Font(bold=True)
    lob_header.alignment = _CENTER_ALIGN

    for c_idx, row_key in enumerate(ROW_ORDER, 2):
        cell = ws.cell(row=h_row, column=c_idx, value=ROW_DISPLAY_NAMES.get(row_key, row_key))
        cell.fill = _META_FILL
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows: one per LOB
    for r_idx, lob in enumerate(active_lobs):
        ws_row = h_row + 1 + r_idx
        lob_cell = ws.cell(row=ws_row, column=1, value=LOB_DISPLAY_NAMES.get(lob, lob))
        lob_cell.font = Font(bold=False)

        lob_data = period_data.data.get(lob, {})
        for c_idx, row_key in enumerate(ROW_ORDER, 2):
            val = lob_data.get(row_key, {}).get(p_key)
            cell = ws.cell(row=ws_row, column=c_idx, value=val)
            cell.number_format = NUMBER_FORMAT
            if (lob, row_key) in period_data.low_confidence_cells:
                cell.fill = _YELLOW_FILL

def _write_meta_sheet(ws, extractions: List[CompanyExtract], stats: Dict[str, Any]):
    """Writes the _meta sheet (Section 11.4)."""
    companies = sorted(list(set(e.company_name for e in extractions)))
    quarters = sorted(list(set(f"{e.quarter}_{e.year}" for e in extractions)))
    
    data = [
        ["Key", "Value"],
        ["extraction_date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["extractor_version", EXTRACTOR_VERSION],
        ["files_processed", stats.get("files_processed", 0)],
        ["files_succeeded", stats.get("files_succeeded", 0)],
        ["files_failed", stats.get("files_failed", 0)],
        ["files_uncategorised", stats.get("files_uncategorised", 0)],
        ["companies", ", ".join(companies)],
        ["quarters", ", ".join(quarters)],
    ]
    
    for r_idx, row in enumerate(data, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = _HEADER_FONT
                cell.fill = _HEADER_FILL
            else:
                cell.fill = _META_FILL

def _sheet_name_for(extract: CompanyExtract) -> str:
    """Build the verification sheet name for an extract (max 31 chars)."""
    name = f"{company_key_to_pascal(extract.company_key)}_{extract.quarter}_{extract.year}"
    return name[:31]


def save_workbook(extractions: List[CompanyExtract], output_path: str, stats: Optional[Dict[str, Any]] = None, year_selection: str = "both"):
    """Create or update the Excel workbook, preserving data from previous runs."""
    if stats is None:
        stats = {}

    output_file = Path(output_path)
    existing_rows = []  # Master_Data rows from previous runs to preserve

    if output_file.exists():
        from openpyxl import load_workbook as _load_wb
        wb = _load_wb(output_path)

        # Files being replaced this run
        new_files = {e.source_file for e in extractions}

        # Read existing Master_Data rows, keep ones NOT being replaced
        if "Master_Data" in wb.sheetnames:
            ws_old = wb["Master_Data"]
            headers = [cell.value for cell in ws_old[1]]

            # Schema guard: only preserve rows if old headers match current structure
            if headers[:len(MASTER_COLUMNS)] == MASTER_COLUMNS:
                try:
                    sf_idx = headers.index("Source_File")
                except ValueError:
                    sf_idx = None

                if sf_idx is not None:
                    for row in ws_old.iter_rows(min_row=2, values_only=True):
                        if row[sf_idx] is None:
                            continue
                        if row[sf_idx] not in new_files:
                            existing_rows.append(list(row))
            else:
                logger.warning(
                    "Existing Master_Data has different column layout — "
                    "discarding old rows and regenerating."
                )

            del wb["Master_Data"]

        # Remove verification sheets being replaced
        for extract in extractions:
            sn = _sheet_name_for(extract)
            if sn in wb.sheetnames:
                del wb[sn]

        # Remove _meta (will be recreated)
        if "_meta" in wb.sheetnames:
            del wb["_meta"]
    else:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default empty sheet

    # 1. Master_Data (always first sheet)
    ws_master = wb.create_sheet("Master_Data", 0)
    _write_master_data(ws_master, extractions, existing_rows=existing_rows, year_selection=year_selection)

    # 2. Individual Verification Sheets (new extractions only; old ones preserved)
    for extract in extractions:
        ws = wb.create_sheet(title=_sheet_name_for(extract))
        _write_verification_sheet(ws, extract)

    # 3. _meta (at the end)
    ws_meta = wb.create_sheet(title="_meta")
    _write_meta_sheet(ws_meta, extractions, stats)

    wb.save(output_path)
    logger.info(f"Excel workbook saved to {output_path}")

def write_validation_summary_sheet(report_path: str, master_path: str, force_company: str = None):
    """
    Reads validation_report.csv and appends a 'Validation_Summary' sheet to master_path.
    Granularity: 1 row per (Company, Quarter, Year).
    """
    import pandas as pd
    df = pd.read_csv(report_path)

    # Group by the specified granularity
    summary = df.pivot_table(
        index=['company', 'quarter', 'year'],
        columns='status',
        aggfunc='size',
        fill_value=0
    ).reset_index()

    # Ensure PASS, WARN, FAIL exist
    for col in ['PASS', 'WARN', 'FAIL', 'SKIP']: # Added SKIP here as it's a valid status
        if col not in summary.columns:
            summary[col] = 0

    # Assumes one PDF per company+quarter+year — will need revision if amended filings are introduced.
    summary['Files_Processed'] = 1

    # Enforce exact column order and rename
    summary = summary.rename(columns={'company': 'Company', 'quarter': 'Quarter', 'year': 'Year'})
    cols = ['Company', 'Quarter', 'Year', 'Files_Processed', 'PASS', 'SKIP', 'WARN', 'FAIL']

    # Calculate Total_Checks from the sum of other status columns
    summary['Total_Checks'] = summary[['PASS', 'SKIP', 'WARN', 'FAIL']].sum(axis=1)

    # Reorder columns to include Total_Checks in the desired position
    cols.insert(4, 'Total_Checks') # Insert 'Total_Checks' after 'Files_Processed'

    summary = summary[cols]
    if force_company:
        try:
            existing = pd.read_excel(master_path, sheet_name="Validation_Summary")
            companies_in_new = set(summary["Company"].unique())
            existing = existing[~existing["Company"].isin(companies_in_new)]
            summary = pd.concat([existing, summary], ignore_index=True)
        except Exception:
            pass
    # Append to Excel
    with pd.ExcelWriter(master_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        summary.to_excel(writer, sheet_name='Validation_Summary', index=False)

def write_validation_detail_sheet(report_path: str, master_path: str, force_company: str = None):
    """
    Reads validation_report.csv, filters for FAILs, and appends a 'Validation_Detail'
    sheet to master_path with Fail_Type and row coloring.
    """
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill

    df = pd.read_csv(report_path)

    # Filter FAILs and WARNs
    cols_map = {
        'company': 'Company', 'quarter': 'Quarter', 'year': 'Year',
        'lob': 'LOB', 'period': 'Period', 'check_name': 'Check_Name',
        'status': 'Status',
        'expected': 'Expected', 'actual': 'Actual', 'delta': 'Delta', 'note': 'Note'
    }

    detail = df[df["status"].isin(["FAIL", "WARN"])].copy()
    if detail.empty:
        logger.info("No failures or warnings found — Validation_Detail sheet written with headers only.")
        detail = pd.DataFrame(columns=list(cols_map.values()))
    else:
        detail = detail.rename(columns=cols_map)[list(cols_map.values())]
        detail = detail.sort_values(by='Status', ascending=True).reset_index(drop=True)

    if force_company:
        try:
            run_companies = set(pd.read_csv(report_path)["company"].unique())
            existing_detail = pd.read_excel(master_path, sheet_name="Validation_Detail")
            if "Company" in existing_detail.columns:
                existing_detail = existing_detail[~existing_detail["Company"].isin(run_companies)]
            detail = pd.concat([existing_detail, detail], ignore_index=True)
        except Exception:
            pass
    # Append to Excel
    with pd.ExcelWriter(master_path, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
        detail.to_excel(writer, sheet_name='Validation_Detail', index=False)

    # Re-open with openpyxl to apply coloring
    wb = load_workbook(master_path)
    ws = wb['Validation_Detail']

    red_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    # Status column index (1-based, no pandas index col since index=False)
    status_col = list(cols_map.values()).index('Status') + 1

    # Iterate through rows (skipping header)
    for row_idx in range(2, ws.max_row + 1):
        status_val = ws.cell(row=row_idx, column=status_col).value
        fill = red_fill if status_val == "FAIL" else yellow_fill
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill

    wb.save(master_path)
    logger.info(f"Validation_Detail sheet written to {master_path}")


def append_to_master(
    extract,
    validation_results: list,
    master_path: str,
) -> int:
    """
    Append extraction results to an existing master Excel file.
    Creates the file if it does not exist.

    Rows are identified by (company_key, quarter, year_code) composite key.
    - Existing rows with source='extractor' are replaced with new data.
    - Existing rows with source='manual' are left untouched.
    - New rows are appended.

    Returns the number of rows written.
    """
    import pandas as pd
    from pathlib import Path as _Path

    master_path = _Path(master_path)
    master_path.parent.mkdir(parents=True, exist_ok=True)

    # Build new rows from this extract
    new_rows = _build_rows(extract, validation_results, source="extractor")

    if not master_path.exists():
        # Create new file
        df_new = pd.DataFrame(new_rows)
        with pd.ExcelWriter(str(master_path), engine="openpyxl") as writer:
            df_new.to_excel(writer, sheet_name="Master_Data", index=False)
        return len(new_rows)

    # Load existing
    try:
        df_existing = pd.read_excel(str(master_path), sheet_name="Master_Data")
    except Exception:
        df_existing = pd.DataFrame()

    if df_existing.empty:
        df_combined = pd.DataFrame(new_rows)
    else:
        # Add source column if not present (backwards compatibility)
        if "source" not in df_existing.columns:
            df_existing["source"] = "extractor"

        df_new = pd.DataFrame(new_rows)
        composite_key = ["company_key", "quarter", "year_code", "lob", "metric"]

        # Build set of new keys for fast lookup
        new_keys = set()
        for row in new_rows:
            key = tuple(row.get(k, "") for k in composite_key)
            new_keys.add(key)

        # Identify which existing rows are protected (manual)
        manual_mask = df_existing["source"] == "manual"

        # Keep manual rows and non-matching rows
        df_keep = df_existing[
            manual_mask | ~df_existing.apply(
                lambda r: tuple(r.get(k, "") for k in composite_key) in new_keys,
                axis=1
            )
        ]

        df_combined = pd.concat([df_keep, df_new], ignore_index=True)

    with pd.ExcelWriter(str(master_path), engine="openpyxl") as writer:
        df_combined.to_excel(writer, sheet_name="Master_Data", index=False)

    return len(new_rows)


def _build_rows(extract, validation_results: list, source: str = "extractor") -> list:
    """
    Convert a CompanyExtract into a list of row dicts for the master sheet.
    Mirrors the structure of the existing Master_Data tab.
    """
    rows = []
    period_data = extract.current_year
    if not period_data:
        return rows
    for lob, metrics in period_data.data.items():
        for metric, values in metrics.items():
            rows.append({
                "source_file":   extract.source_file,
                "company_key":   extract.company_key,
                "company_name":  extract.company_name,
                "form_type":     extract.form_type,
                "fiscal_year":   "",
                "quarter":       extract.quarter,
                "year_code":     extract.year,
                "period":        "current",
                "lob":           lob,
                "metric":        metric,
                "qtr_value":     values.get("qtr"),
                "ytd_value":     values.get("ytd"),
                "source":        source,
            })
    return rows
